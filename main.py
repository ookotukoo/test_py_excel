import os
import requests
import datetime
from bs4 import BeautifulSoup
import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

keys_input = {'d1': '',
              'd1mindate': '20091102',
              'd1maxdate': '',
              'd1day': '1',
              'd1month': '',
              'd1year': '',
              'd2': '',
              'd2mindate': '20091102',
              'd2maxdate': '',
              'd2day': '',
              'd2month': '',
              'd2year': '',
              'bSubmit': 'Показать',
              'pge': '1'}

### блок формирования периода даты за прошлый месяц
# получение первой даты
today = datetime.date.today()
first = today.replace(day=1)

keys_input['d1month'] = (first - datetime.timedelta(days=1)).strftime("%m")
keys_input['d1year'] = (first - datetime.timedelta(days=1)).strftime("%Y")

keys_input['d1'] = keys_input['d1year'] + keys_input['d1month'] + '0' + keys_input['d1day']

# получение 2 даты
today = datetime.date.today()
first = today.replace(day=1)

keys_input['d2day'] = (first - datetime.timedelta(days=1)).strftime("%d")
keys_input['d2month'] = (first - datetime.timedelta(days=1)).strftime("%m")
keys_input['d2year'] = (first - datetime.timedelta(days=1)).strftime("%Y")

keys_input['d2'] = keys_input['d2year'] + keys_input['d2month'] + keys_input['d2day']

### блок формирования периода даты за прошлый месяц закончен

datas_for_excel = []

# 'd_cols': ['Дата JPY/RUB'], 'e_cols': ['Курс JPY/RUB'], 'f_cols': ['Время JPY/RUB'], 'g_cols': ['Результат']

USD_RUB = {'currency': 'USD_RUB'}

JPY_RUB = {'currency': 'JPY_RUB'}

url = 'https://www.moex.com/ru/derivatives/currency-rate.aspx'


def parsing_USD_RUB():
    response_currency = requests.post(requests.get(url, params=USD_RUB).url, data=keys_input).text

    html_code = BeautifulSoup(response_currency, features='lxml')

    all_table = html_code.find('table', class_='tablels')

    all_line_in_table = all_table.find_all('tr')

    rows_all_table = 1
    cell_count = 0
    rows_count = 0
    give_me_fun = 0

    # разбиваем таблицу на строки
    for line in all_line_in_table:

        # через цикл отбрасываю 2 ненужные строки
        if rows_all_table < 3:
            rows_all_table += 1
        else:
            # разбиваем строки на ячейки
            rows_line_in_table = line.find_all('td')

            for cell_in_rows in rows_line_in_table:

                if cell_count == 0:
                    datas_for_excel.append({'Дата USD/RUB': None, 'Курс USD/RUB': None, 'Время USD/RUB': None,
                                            'Дата JPY_RUB': None, 'Курс JPY_RUB': None, 'Время JPY_RUB': None,
                                            'Результат': None})
                    datas_for_excel[give_me_fun]['Дата USD/RUB'] = cell_in_rows.text
                elif cell_count == 3:
                    datas_for_excel[give_me_fun]['Курс USD/RUB'] = float(cell_in_rows.text.replace(',', '.'))
                elif cell_count == 4:
                    datas_for_excel[give_me_fun]['Время USD/RUB'] = cell_in_rows.text
                else:
                    pass
                cell_count += 1
            cell_count = 0
            give_me_fun += 1
        rows_count += 1


def parsing_JPY_RUB():
    response_currency = requests.post(requests.get(url, params=JPY_RUB).url, data=keys_input).text

    html_code = BeautifulSoup(response_currency, features='lxml')

    all_table = html_code.find('table', class_='tablels')

    all_line_in_table = all_table.find_all('tr')

    rows_all_table = 1
    cell_count = 0
    rows_count = 0
    give_me_fun = 0

    # разбиваем таблицу на строки
    for line in all_line_in_table:

        # через цикл отбрасываю 2 ненужные строки
        if rows_all_table < 3:
            rows_all_table += 1
        else:
            # разбиваем строки на ячейки
            rows_line_in_table = line.find_all('td')

            for cell_in_rows in rows_line_in_table:

                if cell_count == 0:
                    datas_for_excel[give_me_fun]['Дата JPY_RUB'] = cell_in_rows.text
                elif cell_count == 3:
                    datas_for_excel[give_me_fun]['Курс JPY_RUB'] = float(cell_in_rows.text.replace(',', '.'))
                elif cell_count == 4:
                    datas_for_excel[give_me_fun]['Время JPY_RUB'] = cell_in_rows.text
                    datas_for_excel[give_me_fun]['Результат'] = datas_for_excel[give_me_fun]['Курс USD/RUB'] / \
                                                                datas_for_excel[give_me_fun]['Курс JPY_RUB']
                else:
                    pass
                cell_count += 1
            cell_count = 0
            give_me_fun += 1
        rows_count += 1


def write_text():

    path_wb = os.path.abspath('Export_one.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.delete_rows(1, ws.max_row)

    ws['A1'].value = 'Дата USD/RUB'
    ws['B1'].value = 'Курс USD/RUB'
    ws['C1'].value = 'Время USD/RUB'
    ws['D1'].value = 'Дата JPY_RUB'
    ws['E1'].value = 'Курс JPY_RUB'
    ws['F1'].value = 'Время JPY_RUB'
    ws['G1'].value = 'Результат'

    row_count = 2

    for string_from_array in datas_for_excel:
        ws.append([string_from_array['Дата USD/RUB'], string_from_array['Курс USD/RUB'],
                   string_from_array['Время USD/RUB'], string_from_array['Дата JPY_RUB'],
                   string_from_array['Курс JPY_RUB'], string_from_array['Время JPY_RUB'],
                   string_from_array['Результат']])

        column_B = 'B' + str(row_count)
        ws[column_B].number_format = '_-* # ##0.00 ₽_-;-* # ##0.00 ₽_-;_-* "-"?? ₽_-;_-@_-'
        column_E = 'E' + str(row_count)
        ws[column_E].number_format = '_-* # ##0.00 ₽_-;-* # ##0.00 ₽_-;_-* "-"?? ₽_-;_-@_-'

        row_count += 1

    dims = {}

    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value + 2

    global max_column

    max_column = ws.max_row

    for colunm_key in dims.keys():
        cell_for_sum = colunm_key + str(max_column + 1)

        sum_formula = '=SUM(' + colunm_key + '2:' + colunm_key + str(max_column) + ')'

        ws[cell_for_sum] = sum_formula

    wb.save(path_wb)
    wb.close()


def send_file():
    login_email = 'send.robot.dadada@gmail.com'

    password_email = os.getenv("PASSWORD")

    to_email = input('Введите почту, куда нужно отправить письмо (тлоько Gmail): ')

    key_strok = {'1': 'у', '2': 'и', '3': 'и', '4': 'и'}

    if 11 <= max_column <= 19:
        strok = ' строк'
    elif str(max_column)[-1] == 0:
        strok = ' строк'
    elif str(max_column)[-1] in key_strok:
        strok = ' строк' + key_strok[str(max_column)[-1]]
    else:
        strok = ' строк'

    message = 'Excel файл содержит: ' + str(max_column) + strok

    file = 'Export_one.xlsx'

    msg = MIMEMultipart()
    msg['From'] = login_email
    msg['To'] = to_email
    msg['Subject'] = 'Курс валют USD_RUB и JPY_RUB за отчётный месяц'
    msg.attach(MIMEText(message))

    fp = open(file, 'rb')
    part = MIMEBase('application', 'vnd.ms-excel')
    part.set_payload(fp.read())
    fp.close()
    encoders.encode_base64(part)

    part.add_header('Content-Disposition', 'attachment', filename=file)
    msg.attach(part)

    smtp = smtplib.SMTP('smtp.gmail.com', 587)
    smtp.ehlo()
    smtp.starttls()
    smtp.login(login_email, password_email)
    smtp.sendmail(login_email, to_email, msg.as_string())
    smtp.quit()


if __name__ == '__main__':
    parsing_USD_RUB()

    parsing_JPY_RUB()

    write_text()

    send_file()
